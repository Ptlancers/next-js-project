import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Protection
import pandas as pd
from .settings import settings  # Assuming settings.py is in the same directory
from .utils import get_month_and_year, replace_dashes_with_space


def create_excel_file(filepath: str):
    wb = Workbook()
    wb.security.lockStructure = True  # Locks the structure of the workbook
    wb.security.lockWindows = True  # Locks the position of the workbook window
    wb.security.lockRevision = True  # Locks revision tracking in the workbook
    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.password = settings.password
    wb.save(filepath)


def create_excel_sheet(wb: Workbook, sheet_name: str, titles: list[str]):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(titles)


def append_data_into_excel(data: dict, section_code: str) -> None:
    month, year = get_month_and_year(data["date"])
    file_path = os.path.join(
        settings.excelFolderPath, f"{section_code}-receipt-{year}.xlsx"
    )

    if not os.path.exists(file_path):
        create_excel_file(file_path)

    wb = load_workbook(file_path)
    sheet_name: str = month
    titles: list[str] = ["Serial Number"] + list(replace_dashes_with_space(data).keys())
    create_excel_sheet(wb, sheet_name, titles)

    ws = wb[sheet_name]
    data_with_serial = {"Serial Number": len(ws["A"])}
    data_with_serial.update(data)

    df = pd.DataFrame([data_with_serial])

    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    for row in ws.iter_rows(
        min_row=len(ws["A"]) + 1, max_row=len(ws["A"]) + len(df), max_col=len(titles)
    ):
        for cell in row:
            cell.protection = Protection(locked=True)

    wb.save(file_path)


def update_data_into_excel(data: dict, section_code: str, receipt_number: str) -> None:
    month, year = get_month_and_year(data.get("date"))
    file_path = os.path.join(
        settings.excelFolderPath, f"{section_code}-receipt-{year}.xlsx"
    )

    if not os.path.exists(file_path):
        create_excel_file(file_path)

    wb = load_workbook(file_path)
    sheet_name: str = month
    ws = wb[sheet_name]
    data = replace_dashes_with_space(data)

    df = pd.DataFrame([data])

    for r in dataframe_to_rows(df, index=False, header=False):
        data = r
        # print(r)

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column
    ):
        if row[2].value == receipt_number:
            for index in range(len(data)):
                cell = row[index]
                value = data[index]
                cell.value = value
            for cell in row:
                cell.protection = Protection(locked=True)

            wb.save(file_path)
            return


# temp_data={
#     "date": '2024-04-27',
#     "donor_registration_number": '7395892580',
#     "receipt_number": '90',
#     "donor_name": 'vijayaKumar',
#     "unique_identification_number": '1234567890',
#     "address": 's,dncasddfnkwenrwu',
#     "donation_type": 'corpus',
#     "mode_of_receipt": 'digital',
#     "transaction_id": '1234567890',
#     "donated_amount": '501',
#     "donated_amount_letters": 'five hundred one',
#     "section_code": 'Section80',
#     "unique_registration_number": '1234567890',
#     "date_insurance_of_urn": '2024-04-05'
# }
# append_data_into_excel(temp_data,temp_data["section_code"])
